"""Scrittura Excel / CSV dei pacchetti ANT+ grezzi."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font

COLUMNS: Sequence[str] = (
    "timestamp_iso",
    "elapsed_s",
    "device_id",
    "page",
    "page_name",
    "instantaneous_power_w",
    "average_power_w",
    "cadence_rpm",
    "left_power_w",
    "right_power_w",
    "torque_nm",
    "angular_velocity_rad_s",
    "raw_bytes_hex",
)


def _row_values(row: dict[str, Any]) -> list[Any]:
    return [row.get(col) for col in COLUMNS]


def write_xlsx(path: Path, rows: Iterable[dict[str, Any]], meta: dict[str, Any] | None = None) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "raw"

    header_font = Font(bold=True)
    ws.append(list(COLUMNS))
    for cell in ws[1]:
        cell.font = header_font

    count = 0
    for row in rows:
        ws.append(_row_values(row))
        count += 1

    for col in ws.columns:
        letter = col[0].column_letter
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col[:50])
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 42)

    meta_ws = wb.create_sheet("meta")
    meta_ws.append(["chiave", "valore"])
    meta_ws["A1"].font = header_font
    meta_ws["B1"].font = header_font
    payload = {
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "rows": count,
        "note": "Dati raw ANT+ Bicycle Power dalla Wattbike (o mock). Nessuna elaborazione applicata.",
        **(meta or {}),
    }
    for key, value in payload.items():
        meta_ws.append([key, value])

    wb.save(path)
    return path


def write_csv(path: Path, rows: Iterable[dict[str, Any]]) -> Path:
    import csv

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(COLUMNS), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col) for col in COLUMNS})
    return path
